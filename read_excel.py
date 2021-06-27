from openpyxl import load_workbook

# to read xlsx files we need openpyxl
wb = load_workbook(filename = 'sampledatainsurance.xlsx')
print("The number of worksheets is {0}".format(len(wb.sheetnames)))
print("Names of worksheets is {0}".format(wb.sheetnames))
wb_sheets = wb.sheetnames
# sheet_obj = wb.active
wb1 = wb[wb_sheets[0]]
 

for sheet in wb:
    # print(sheet.title)
    wb1 = wb[sheet.title]
    print(sheet.title, '...number of rows..', wb1.max_row, '..columns..', wb1.max_column)
    # print colum names 
    for i in range (1,wb1.max_column+1):
        cell_obj = wb1.cell(row = 1, column = i)
        print(cell_obj.value)

wb1 = wb['PolicyData']
for i in range(1,wb1.max_row+1):
    for j in range(1,2):
        # for j in range(1,wb1.max_column+1):
        cell_obj = wb1.cell(row = i, column = j)
        print(cell_obj.value)
